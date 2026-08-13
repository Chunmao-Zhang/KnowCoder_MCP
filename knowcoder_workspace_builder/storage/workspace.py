"""Build, validate, and read the executable knowledge Workspace."""

from __future__ import annotations

import hashlib
import json
import os
import shutil
from pathlib import Path, PurePosixPath
from typing import Any
from uuid import uuid4

from knowcoder_workspace_builder.contracts.errors import ContractError, MissingStateError
from knowcoder_workspace_builder.contracts.workspace import PUBLIC_WORKSPACE_DIRECTORIES, PUBLIC_WORKSPACE_FILES

from .audit import build_audit
from .instances import validate_instances
from .paths import SessionPaths, is_within
from .readme import validate_workspace_readme
from .schema import ParsedSchema, parse_schema
from .transaction import AtomicWriter, read_json


LOADER_SOURCE = '''"""Instance-format loader for this executable knowledge Workspace."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class WorkspaceGraph:
    root: Path
    schema: dict[str, Any]
    manifest: dict[str, Any]
    entities: tuple[dict[str, Any], ...]
    relations: tuple[dict[str, Any], ...]
    source_chunks: tuple[dict[str, Any], ...]
    entity_index: dict[tuple[str, str], dict[str, Any]]

    def entity(self, entity_type: str, entity_id: str) -> dict[str, Any]:
        key = (entity_type, str(entity_id))
        if key not in self.entity_index:
            raise KeyError(f"Unknown entity: {entity_type}:{entity_id}")
        return self.entity_index[key]

    def outgoing(self, entity_type: str, entity_id: str) -> tuple[dict[str, Any], ...]:
        key = (entity_type, str(entity_id))
        return tuple(
            relation
            for relation in self.relations
            if (relation["head"]["type"], str(relation["head"]["id"])) == key
        )


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        value = json.loads(line)
        if not isinstance(value, dict):
            raise ValueError(f"{path.name} line {number} must be an object")
        records.append(value)
    return records


def load_workspace(root: str | Path) -> WorkspaceGraph:
    workspace = Path(root).expanduser().resolve(strict=True)
    schema = json.loads((workspace / "ontology/schema.json").read_text(encoding="utf-8"))
    manifest = json.loads((workspace / "data/manifest.json").read_text(encoding="utf-8"))
    entities = _read_jsonl(workspace / "data/entities.jsonl")
    relations = _read_jsonl(workspace / "data/relations.jsonl")
    source_chunks = _read_jsonl(workspace / "data/source_chunks.jsonl")
    definitions = schema.get("entities")
    if not isinstance(definitions, list):
        raise ValueError("ontology/schema.json requires entity definitions")
    source_records = manifest.get("sources")
    if not isinstance(source_records, list):
        raise ValueError("data/manifest.json requires a sources list")
    source_ids = {str(item.get("source_id") or "") for item in source_records if isinstance(item, dict)}
    chunk_ref_index: set[tuple[str, str]] = set()
    for record in source_chunks:
        source_id = str(record.get("source_id") or "").strip()
        chunk_id = str(record.get("chunk_id") or "").strip()
        key = (source_id, chunk_id)
        if not source_id or source_id not in source_ids or not chunk_id or key in chunk_ref_index:
            raise ValueError(f"Invalid source chunk record: {source_id}:{chunk_id}")
        chunk_ref_index.add(key)
    entity_index: dict[tuple[str, str], dict[str, Any]] = {}
    for record in entities:
        entity_type = record.get("type")
        entity_id = record.get("id")
        name = record.get("name")
        if not isinstance(entity_type, str) or not entity_type.strip():
            raise ValueError("Entity type must be non-empty text")
        if not ((isinstance(entity_id, str) and entity_id.strip()) or (isinstance(entity_id, int) and not isinstance(entity_id, bool))):
            raise ValueError(f"Entity ID is invalid: {entity_type}")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"Entity name is invalid: {entity_type}:{entity_id}")
        key = (entity_type.strip(), str(entity_id))
        if key in entity_index:
            raise ValueError(f"Missing or duplicate entity ID: {key}")
        attributes = record.get("attributes")
        if not isinstance(attributes, dict):
            raise ValueError(f"Entity attributes must be an object: {key}")
        refs = record.get("source_refs")
        if not isinstance(refs, list) or any(not isinstance(ref, str) or not ref.strip() for ref in refs):
            raise ValueError(f"Entity source_refs are invalid: {key}")
        if len(refs) != len(set(refs)) or set(refs) - source_ids:
            raise ValueError(f"Entity source_refs are invalid: {key}")
        evidence_refs = record.get("evidence_refs", [])
        if not isinstance(evidence_refs, list):
            raise ValueError(f"Entity evidence_refs are invalid: {key}")
        for ref in evidence_refs:
            if not isinstance(ref, dict):
                raise ValueError(f"Entity evidence_refs are invalid: {key}")
            ref_key = (str(ref.get("source_id") or ""), str(ref.get("chunk_id") or ""))
            if ref_key[0] not in refs or ref_key not in chunk_ref_index:
                raise ValueError(f"Entity evidence_refs are invalid: {key}")
        entity_index[key] = record
    for relation in relations:
        relation_type = relation.get("type")
        if not isinstance(relation_type, str) or not relation_type.strip():
            raise ValueError("Relation type must be non-empty text")
        head = relation.get("head")
        tail = relation.get("tail")
        if not isinstance(head, dict) or not isinstance(tail, dict):
            raise ValueError(f"Relation endpoints are invalid: {relation_type}")
        head_key = (str(head.get("type") or "").strip(), str(head.get("id") if head.get("id") is not None else ""))
        tail_key = (str(tail.get("type") or "").strip(), str(tail.get("id") if tail.get("id") is not None else ""))
        if not all((*head_key, *tail_key)):
            raise ValueError(f"Relation endpoint values are invalid: {relation_type}")
        if head_key not in entity_index or tail_key not in entity_index:
            raise ValueError(f"Relation points to a missing entity: {relation_type}")
        attributes = relation.get("attributes")
        if not isinstance(attributes, dict):
            raise ValueError(f"Relation attributes must be an object: {relation_type}")
        refs = relation.get("source_refs")
        if not isinstance(refs, list) or any(not isinstance(ref, str) or not ref.strip() for ref in refs):
            raise ValueError(f"Relation source_refs are invalid: {relation_type}")
        if len(refs) != len(set(refs)) or set(refs) - source_ids:
            raise ValueError(f"Relation source_refs are invalid: {relation_type}")
        evidence_refs = relation.get("evidence_refs", [])
        if not isinstance(evidence_refs, list):
            raise ValueError(f"Relation evidence_refs are invalid: {relation_type}")
        for ref in evidence_refs:
            if not isinstance(ref, dict):
                raise ValueError(f"Relation evidence_refs are invalid: {relation_type}")
            ref_key = (str(ref.get("source_id") or ""), str(ref.get("chunk_id") or ""))
            if ref_key[0] not in refs or ref_key not in chunk_ref_index:
                raise ValueError(f"Relation evidence_refs are invalid: {relation_type}")
    expected = manifest.get("records") or {}
    if expected.get("entities") != len(entities) or expected.get("relations") != len(relations):
        raise ValueError("Manifest record counts do not match instance files")
    return WorkspaceGraph(
        root=workspace,
        schema=schema,
        manifest=manifest,
        entities=tuple(entities),
        relations=tuple(relations),
        source_chunks=tuple(source_chunks),
        entity_index=entity_index,
    )
'''


def _jsonl(records: list[dict[str, Any]]) -> str:
    return "".join(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n" for record in records)


def _source_path(paths: SessionPaths, value: object) -> Path:
    text = str(value or "").strip()
    if not text:
        raise ContractError("Workspace source record requires file_path")
    prefix = "/.knowcoder_workspace/"
    candidate = paths.root / text.removeprefix(prefix) if text.startswith(prefix) else Path(text).expanduser()
    resolved = candidate.resolve(strict=True)
    if not is_within(resolved, paths.root.resolve(strict=True)):
        raise ContractError("Workspace source file is outside the current Session", path=str(resolved))
    if not resolved.is_file():
        raise ContractError("Workspace source path is not a file", path=str(resolved))
    return resolved


def _source_body(path: Path, title: str) -> str:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - declared runtime dependency
            raise ContractError("Excel source rendering requires openpyxl") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sections = [f"# {title}", ""]
            rendered_sheet = False
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                rows = [row for row in rows if any(value is not None and str(value).strip() for value in row)]
                if not rows:
                    continue
                rendered_sheet = True
                width = max(len(row) for row in rows)

                def cells(row: list[Any]) -> list[str]:
                    padded = [*row, *([None] * (width - len(row)))]
                    values: list[str] = []
                    for value in padded:
                        if value is None:
                            text = ""
                        elif isinstance(value, float) and value.is_integer():
                            text = str(int(value))
                        else:
                            text = str(value).strip()
                        values.append(text.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>"))
                    return values

                header = cells(rows[0])
                header = [value or f"Column {index}" for index, value in enumerate(header, start=1)]
                sections.extend(
                    [
                        f"## Sheet: {sheet.title}",
                        "",
                        "| " + " | ".join(header) + " |",
                        "| " + " | ".join("---" for _ in header) + " |",
                    ]
                )
                sections.extend("| " + " | ".join(cells(row)) + " |" for row in rows[1:])
                sections.append("")
            if not rendered_sheet:
                raise ContractError("Excel source has no readable rows", path=str(path))
            return "\n".join(sections).rstrip() + "\n"
        finally:
            workbook.close()
    if path.suffix.casefold() == ".json":
        value = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(value, dict) and isinstance(value.get("results"), list):
            lines = [f"# {title}", "", f"Query: {value.get('query') or ''}", "", "## Results", ""]
            for item in value["results"]:
                if not isinstance(item, dict):
                    continue
                lines.extend(
                    [
                        f"### {item.get('title') or 'Untitled source'}",
                        "",
                        f"URL: {item.get('link') or item.get('url') or ''}",
                        "",
                        str(item.get("snippet") or item.get("content") or "").strip(),
                        "",
                    ]
                )
            return "\n".join(lines).rstrip() + "\n"
        if isinstance(value, dict) and any(key in value for key in ("snippet", "content", "text")):
            return "\n".join(
                [
                    f"# {title}",
                    "",
                    f"URL: {value.get('url') or value.get('link') or ''}",
                    "",
                    str(value.get("snippet") or value.get("content") or value.get("text") or "").strip(),
                    "",
                ]
            )
        rendered = json.dumps(value, ensure_ascii=False, indent=2)
        return f"# {title}\n\n```json\n{rendered}\n```\n"
    text = path.read_text(encoding="utf-8", errors="strict")
    if path.suffix.casefold() == ".md":
        return text.rstrip() + "\n"
    language = path.suffix.casefold().lstrip(".") or "text"
    return f"# {title}\n\n```{language}\n{text.rstrip()}\n```\n"


def _chunk_records(
    paths: SessionPaths,
    source: dict[str, Any],
    source_id: str,
    *,
    content: str,
    workspace_offset: int,
) -> list[dict[str, Any]]:
    chunk_path_value = str(source.get("chunk_path") or "").strip()
    if not chunk_path_value:
        return []
    chunk_path = _source_path(paths, chunk_path_value)
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for line_number, line in enumerate(chunk_path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        value = json.loads(line)
        if not isinstance(value, dict):
            raise ContractError("Source chunk record must be an object", source_id=source_id, line=line_number)
        chunk_id = str(value.get("chunk_id") or "").strip()
        chunk_source_id = str(value.get("source_id") or "").strip()
        if not chunk_id or chunk_source_id != source_id or chunk_id in seen:
            raise ContractError("Source chunk identity is invalid", source_id=source_id, chunk_id=chunk_id)
        start = value.get("start")
        end = value.get("end")
        if not isinstance(start, int) or isinstance(start, bool) or not isinstance(end, int) or isinstance(end, bool):
            raise ContractError("Source chunk offsets must be integers", source_id=source_id, chunk_id=chunk_id)
        if start < 0 or end <= start:
            raise ContractError("Source chunk offsets are invalid", source_id=source_id, chunk_id=chunk_id)
        content_sha256 = str(value.get("content_sha256") or "").strip()
        if not content_sha256:
            raise ContractError("Source chunk requires a content hash", source_id=source_id, chunk_id=chunk_id)
        if end > len(content):
            raise ContractError("Source chunk exceeds its source content", source_id=source_id, chunk_id=chunk_id)
        actual_hash = hashlib.sha256(content[start:end].strip().encode("utf-8")).hexdigest()
        if actual_hash != content_sha256:
            raise ContractError("Source chunk content hash does not match its offsets", source_id=source_id, chunk_id=chunk_id)
        records.append(
            {
                "source_id": source_id,
                "chunk_id": chunk_id,
                "heading": str(value.get("heading") or ""),
                "start": start + workspace_offset,
                "end": end + workspace_offset,
                "content_sha256": content_sha256,
            }
        )
        seen.add(chunk_id)
    return records


def _source_documents(
    paths: SessionPaths,
    sources: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, str], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    documents: dict[str, str] = {}
    source_chunks: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for source in sources:
        if not isinstance(source, dict):
            raise ContractError("Workspace source record must be an object")
        source_id = str(source.get("source_id") or "").strip()
        if not source_id or source_id in seen_ids:
            raise ContractError("Workspace source IDs must be unique non-empty text", source_id=source_id)
        seen_ids.add(source_id)
        path = _source_path(paths, source.get("file_path"))
        content_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        filename = f"{hashlib.sha256(source_id.encode('utf-8')).hexdigest()[:16]}.md"
        title = str(source.get("title") or path.name).strip()
        url = str(source.get("url") or "").strip()
        links = [str(item).strip() for item in source.get("links") or [] if str(item).strip()]
        frontmatter = [
            "---",
            "type: Source",
            f"source_id: {json.dumps(source_id, ensure_ascii=False)}",
            f"title: {json.dumps(title, ensure_ascii=False)}",
            f"source_kind: {json.dumps(str(source.get('source_kind') or 'unknown'), ensure_ascii=False)}",
            f"uri: {json.dumps(url, ensure_ascii=False)}",
            f"content_sha256: {json.dumps(content_hash)}",
            "---",
            "",
        ]
        prefix = "\n".join(frontmatter)
        body = _source_body(path, title)
        documents[filename] = prefix + body
        chunks = _chunk_records(
            paths,
            source,
            source_id,
            content=body,
            workspace_offset=len(prefix),
        )
        source_chunks.extend(chunks)
        records.append(
            {
                "source_id": source_id,
                "title": title,
                "source_kind": str(source.get("source_kind") or "unknown"),
                "url": url,
                "links": links,
                "retrieved_at": source.get("retrieved_at"),
                "status": str(source.get("status") or "active"),
                "superseded_by": str(source.get("superseded_by") or ""),
                "workspace_path": f"data/source/{filename}",
                "content_sha256": content_hash,
                "chunk_count": len(chunks),
            }
        )
    return records, documents, source_chunks


def _workspace_yaml(schema_version: int, data_version: int) -> str:
    return "\n".join(
        [
            "format: executable-knowledge-workspace",
            "format_version: 1",
            'okf_version: "0.1"',
            f"schema_version: {schema_version}",
            f"data_version: {data_version}",
            "ontology: ontology/types.py",
            "loader: ontology/loader.py",
            "schema_contract: ontology/schema.json",
            "entities: data/entities.jsonl",
            "relations: data/relations.jsonl",
            "source_chunks: data/source_chunks.jsonl",
            "manifest: data/manifest.json",
            "sources: data/source/",
            "knowledge: knowledge/",
            "",
        ]
    )


def _ontology_readme(schema: ParsedSchema) -> str:
    entities = "\n".join(f"- `{item.name}`: {item.description}" for item in schema.entities)
    relations = "\n".join(
        f"- `{field.name}`: `{entity.name}` -> `{field.value_type}`. {field.description}"
        for entity in schema.entities
        for field in entity.relations
    )
    return (
        "# Ontology\n\n"
        "`types.py` is the executable schema declaration. `schema.json` is its language-neutral contract. "
        "`loader.py` validates the current data and builds the object graph.\n\n"
        f"## Entities\n\n{entities}\n\n"
        f"## Relations\n\n{relations}\n\n"
        "## Validation\n\n"
        "The loader checks canonical Instance fields, entity IDs, relation endpoints, source references, and manifest counts.\n"
    )


class WorkspaceRepository:
    def __init__(self, paths: SessionPaths) -> None:
        self.paths = paths

    def commit(
        self,
        *,
        schema_source: str,
        instances: dict[str, Any],
        sources: list[dict[str, Any]],
        schema_version: int,
        data_version: int,
        readme: str,
    ) -> dict[str, str]:
        schema = parse_schema(schema_source, require_relations=False)
        source_records, source_documents, source_chunks = _source_documents(self.paths, sources)
        known_sources = {item["source_id"] for item in source_records}
        known_chunk_refs = {
            (str(item["source_id"]), str(item["chunk_id"]))
            for item in source_chunks
        }
        normalized = validate_instances(
            instances,
            schema,
            allowed_source_ids=known_sources,
            allowed_chunk_refs=known_chunk_refs,
        )
        referenced = {
            str(source_id)
            for record in [*normalized["entities"], *normalized["relations"]]
            for source_id in record.get("source_refs") or []
        }
        unknown = sorted(referenced - known_sources)
        if unknown:
            raise ContractError("Workspace instances reference unknown sources", source_ids=unknown)
        audit = build_audit(
            normalized,
            schema,
            source_records,
            schema_version=schema_version,
            data_version=data_version,
        )
        manifest = {
            "format_version": 1,
            "schema_version": schema_version,
            "data_version": data_version,
            "files": {
                "entities": "data/entities.jsonl",
                "relations": "data/relations.jsonl",
                "source_chunks": "data/source_chunks.jsonl",
                "source_directory": "data/source",
            },
            "records": {
                "entities": len(normalized["entities"]),
                "relations": len(normalized["relations"]),
            },
            "entity_types": audit["entity_types"],
            "relation_types": audit["relation_types"],
            "sources": source_records,
        }
        validate_workspace_readme(readme)
        publication_id = f"data-{data_version:08d}-{uuid4().hex[:12]}"
        versions_root = self.paths.intermediate / "workspace_versions"
        version_path = versions_root / publication_id
        current_path = versions_root / "current.json"
        staged = self.paths.intermediate / f".workspace-{uuid4().hex}.tmp"
        backup = self.paths.intermediate / f".workspace-{uuid4().hex}.bak"
        writer = AtomicWriter(self.paths)
        published = False
        try:
            for relative in PUBLIC_WORKSPACE_DIRECTORIES:
                (staged / relative).mkdir(parents=True, exist_ok=True)
            review_directory = self.paths.workspace / "review"
            if review_directory.is_dir():
                shutil.copytree(review_directory, staged / "review")
            writer.text(staged / "README.md", readme)
            writer.text(staged / "workspace.yaml", _workspace_yaml(schema_version, data_version))
            writer.text(staged / "ontology" / "README.md", _ontology_readme(schema))
            writer.text(staged / "ontology" / "types.py", schema_source.rstrip() + "\n")
            writer.text(staged / "ontology" / "loader.py", LOADER_SOURCE)
            writer.json(staged / "ontology" / "schema.json", schema.outline())
            writer.text(staged / "data" / "entities.jsonl", _jsonl(normalized["entities"]))
            writer.text(staged / "data" / "relations.jsonl", _jsonl(normalized["relations"]))
            writer.text(staged / "data" / "source_chunks.jsonl", _jsonl(source_chunks))
            writer.json(staged / "data" / "manifest.json", manifest)
            for filename, content in source_documents.items():
                writer.text(staged / "data" / "source" / filename, content)
            versions_root.mkdir(parents=True, exist_ok=True)
            shutil.copytree(staged, version_path)
            if self.paths.workspace.exists():
                os.replace(self.paths.workspace, backup)
            os.replace(staged, self.paths.workspace)
            published = True
            writer.json(
                current_path,
                {
                    "publication_id": publication_id,
                    "schema_version": schema_version,
                    "data_version": data_version,
                },
            )
            if backup.exists():
                shutil.rmtree(backup, ignore_errors=True)
        except Exception:
            if staged.exists():
                shutil.rmtree(staged)
            if published and self.paths.workspace.exists():
                shutil.rmtree(self.paths.workspace)
            if backup.exists():
                os.replace(backup, self.paths.workspace)
            if version_path.exists():
                shutil.rmtree(version_path)
            raise
        return {name: self.paths.relative_to_project(self.paths.workspace / name) for name in PUBLIC_WORKSPACE_FILES}

    def validate_ready(self) -> dict[str, Any]:
        missing_directories = [
            name for name in PUBLIC_WORKSPACE_DIRECTORIES if not (self.paths.workspace / name).is_dir()
        ]
        missing_files = [name for name in PUBLIC_WORKSPACE_FILES if not (self.paths.workspace / name).is_file()]
        if missing_directories or missing_files:
            raise MissingStateError(
                "Public Workspace is incomplete",
                missing_directories=missing_directories,
                missing_files=missing_files,
            )
        if any((self.paths.workspace / "knowledge").iterdir()):
            raise ContractError("The reserved knowledge directory must remain empty")
        schema_source = (self.paths.workspace / "ontology/types.py").read_text(encoding="utf-8")
        schema = parse_schema(schema_source, require_relations=False)
        schema_contract = read_json(self.paths.workspace / "ontology/schema.json")
        if schema_contract != schema.outline():
            raise ContractError("ontology/schema.json does not match ontology/types.py")
        entities = _read_jsonl(self.paths.workspace / "data/entities.jsonl")
        relations = _read_jsonl(self.paths.workspace / "data/relations.jsonl")
        manifest = read_json(self.paths.workspace / "data/manifest.json")
        if not isinstance(manifest, dict) or not isinstance(manifest.get("sources"), list):
            raise ContractError("data/manifest.json is invalid")
        source_ids = {str(item.get("source_id") or "") for item in manifest["sources"] if isinstance(item, dict)}
        source_chunks = _read_jsonl(self.paths.workspace / "data/source_chunks.jsonl")
        chunk_refs: set[tuple[str, str]] = set()
        for item in source_chunks:
            source_id = str(item.get("source_id") or "").strip()
            chunk_id = str(item.get("chunk_id") or "").strip()
            key = (source_id, chunk_id)
            if not source_id or source_id not in source_ids or not chunk_id or key in chunk_refs:
                raise ContractError("Workspace source chunk map is invalid", source_id=source_id, chunk_id=chunk_id)
            chunk_refs.add(key)
        instances = validate_instances(
            {"entities": entities, "relations": relations},
            schema,
            allowed_source_ids=source_ids,
            allowed_chunk_refs=chunk_refs,
        )
        referenced = {
            str(source_id)
            for record in [*instances["entities"], *instances["relations"]]
            for source_id in record.get("source_refs") or []
        }
        if referenced - source_ids:
            raise ContractError("Workspace contains unresolved source_refs", source_ids=sorted(referenced - source_ids))
        for source in manifest["sources"]:
            if not isinstance(source, dict):
                raise ContractError("Workspace source manifest record must be an object")
            source_path = self.paths.workspace / str(source.get("workspace_path") or "")
            if not source_path.is_file() or source_path.suffix.casefold() != ".md":
                raise ContractError("Workspace source Markdown is missing", source_id=source.get("source_id"))
        expected = manifest.get("records") or {}
        if expected.get("entities") != len(entities) or expected.get("relations") != len(relations):
            raise ContractError("Workspace manifest counts do not match data files")
        readme = (self.paths.workspace / "README.md").read_text(encoding="utf-8")
        validate_workspace_readme(readme)
        return {
            "files": {name: str(self.paths.workspace / name) for name in PUBLIC_WORKSPACE_FILES},
            "schema": schema.outline(),
            "entity_count": len(entities),
            "relation_count": len(relations),
            "manifest": manifest,
        }

    def read_artifact(self, name: str) -> Path:
        if name not in PUBLIC_WORKSPACE_FILES:
            requested = PurePosixPath(str(name))
            is_source_path = (
                not requested.is_absolute()
                and requested.as_posix() == str(name)
                and len(requested.parts) == 3
                and requested.parts[:2] == ("data", "source")
                and requested.suffix.casefold() == ".md"
            )
            manifest_path = self.paths.workspace / "data" / "manifest.json"
            manifest = read_json(manifest_path) if manifest_path.is_file() else {}
            sources = manifest.get("sources") if isinstance(manifest, dict) else None
            allowed_sources = {
                str(item.get("workspace_path") or "")
                for item in sources or []
                if isinstance(item, dict)
            }
            if not is_source_path or str(name) not in allowed_sources:
                raise ContractError(
                    "Unknown public Workspace artifact",
                    artifact=name,
                    allowed=[*PUBLIC_WORKSPACE_FILES, "data/source/<manifest workspace_path>.md"],
                )
        path = self.paths.workspace / name
        if not path.is_file():
            raise MissingStateError("Workspace artifact does not exist", artifact=name)
        return path


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ContractError("Workspace JSONL is invalid", path=str(path), line=number) from exc
        if not isinstance(value, dict):
            raise ContractError("Workspace JSONL line must be an object", path=str(path), line=number)
        records.append(value)
    return records
