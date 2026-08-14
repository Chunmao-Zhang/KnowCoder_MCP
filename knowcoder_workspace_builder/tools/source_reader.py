"""Read registered evidence sources within the current Session."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from langchain_core.tools import tool

from knowcoder_workspace_builder.runtime.invocation_context import (
    active_invocation_context,
)
from knowcoder_workspace_builder.runtime.parallel_units import CONSECUTIVE_FAILURE_LIMIT
from knowcoder_workspace_builder.runtime.session_context import active_session_paths
from knowcoder_workspace_builder.runtime.token_chunks import token_chunks
from knowcoder_workspace_builder.runtime.virtual_paths import virtual_path_for
from knowcoder_workspace_builder.runtime.workspace_sources import source_records
from knowcoder_workspace_builder.storage.transaction import AtomicWriter

from .path_utils import resolve_path
from .web_content import relevant_chunks, relevant_excerpt
from .web_fetch import load_source_chunks, load_web_fetch_settings

MAX_TEXT_CHARS = 200_000
MAX_PREVIEW_ROWS = 50
MAX_BATCH_SOURCES = 40
MAX_BATCH_TEXT_CHARS = 96_000
MAX_CHUNKS_PER_SOURCE = 3
MAX_CHARS_PER_SOURCE = 4_000
MAX_UNSTRUCTURED_SOURCE_BATCH = MAX_BATCH_SOURCES
ASSIGNED_SOURCES_TOKEN = "*"


def _registered_source(path: Path) -> dict[str, Any]:
    paths = active_session_paths()
    for record in source_records(paths.root):
        file_path = str(record.get("file_path") or "")
        if not file_path:
            continue
        try:
            registered_path = resolve_path(file_path)
        except (OSError, ValueError):
            continue
        if registered_path.resolve(strict=False) == path.resolve(strict=False):
            return record
    raise ValueError(f"source is not registered in the current Session: {path.name}")


def _base_record(path: Path, registered: dict[str, Any]) -> dict[str, Any]:
    paths = active_session_paths()
    return {
        "source_id": str(registered["source_id"]),
        "source_kind": str(registered.get("source_kind") or "upload"),
        "file_path": virtual_path_for(paths.root, path),
        "file_type": path.suffix.casefold().lstrip(".") or "txt",
        "title": str(registered.get("title") or path.name),
        "url": str(registered.get("url") or ""),
    }


def _read_tabular(path: Path, registered: dict[str, Any], delimiter: str) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        preview: list[dict[str, str]] = []
        row_count = 0
        for row in reader:
            row_count += 1
            if len(preview) < MAX_PREVIEW_ROWS:
                preview.append(dict(row))
        columns = list(reader.fieldnames or [])
    return {
        **_base_record(path, registered),
        "columns": columns,
        "sample_rows": preview,
        "chunks": [],
        "metadata": {"row_count": row_count, "preview_row_count": len(preview), "size_bytes": path.stat().st_size},
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_excel(path: Path, registered: dict[str, Any]) -> dict[str, Any]:
    """Read the first non-empty Excel sheet as a structured table preview."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency should be installed
        raise ValueError("Excel support requires openpyxl to be installed") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = None
        for candidate in workbook.worksheets:
            if candidate.max_row and candidate.max_row > 0:
                sheet = candidate
                break
        if sheet is None:
            raise ValueError("Excel workbook has no readable sheets")

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ValueError("Excel sheet is empty") from exc

        columns: list[str] = []
        seen: set[str] = set()
        for index, cell in enumerate(header_row, start=1):
            name = _cell_text(cell) or f"column_{index}"
            if name in seen:
                name = f"{name}_{index}"
            seen.add(name)
            columns.append(name)

        preview: list[dict[str, str]] = []
        row_count = 0
        for raw in rows_iter:
            values = [_cell_text(cell) for cell in raw]
            if not any(values):
                continue
            row_count += 1
            if len(preview) < MAX_PREVIEW_ROWS:
                padded = values + [""] * max(0, len(columns) - len(values))
                preview.append({columns[index]: padded[index] for index in range(len(columns))})
        return {
            **_base_record(path, registered),
            "columns": columns,
            "sample_rows": preview,
            "chunks": [],
            "metadata": {
                "row_count": row_count,
                "preview_row_count": len(preview),
                "size_bytes": path.stat().st_size,
                "sheet_name": sheet.title,
            },
        }
    finally:
        workbook.close()


def _read_json(path: Path, registered: dict[str, Any]) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    records = value if isinstance(value, list) else [value]
    if not all(isinstance(item, dict) for item in records):
        text = json.dumps(value, ensure_ascii=False)
        return _text_record(path, registered, text)
    columns = list(dict.fromkeys(key for item in records for key in item))
    preview = [dict(item) for item in records[:MAX_PREVIEW_ROWS]]
    source_kind = str(registered.get("source_kind") or "")
    if source_kind == "web_search_bundle":
        return _text_record(path, registered, json.dumps(value, ensure_ascii=False, indent=2))
    if source_kind in {"web", "web_crawl"}:
        text = "\n\n".join(
            str(item.get("snippet") or item.get("content") or item.get("text") or json.dumps(item, ensure_ascii=False))
            for item in records
        )
        return _text_record(path, registered, text)
    return {
        **_base_record(path, registered),
        "columns": columns,
        "sample_rows": preview,
        "chunks": [],
        "metadata": {
            "row_count": len(records),
            "preview_row_count": len(preview),
            "size_bytes": path.stat().st_size,
        },
    }


def _text_record(path: Path, registered: dict[str, Any], text: str | None = None) -> dict[str, Any]:
    content = path.read_text(encoding="utf-8", errors="strict") if text is None else text
    if len(content) > MAX_TEXT_CHARS:
        raise ValueError(
            f"text source exceeds the {MAX_TEXT_CHARS} character per-source limit; split it into registered sources"
        )
    settings = load_web_fetch_settings()
    chunks = [
        {
            **item,
            "chunk_id": f"{registered['source_id']}#chunk_{index:04d}",
        }
        for index, item in enumerate(
            token_chunks(
                content,
                target_tokens=settings.schema_chunk_target_tokens,
                overlap_tokens=settings.schema_chunk_overlap_tokens,
            ),
            start=1,
        )
    ]
    return {
        **_base_record(path, registered),
        "columns": [],
        "sample_rows": [],
        "chunks": chunks,
        "metadata": {"character_count": len(content), "chunk_count": len(chunks), "size_bytes": path.stat().st_size},
    }


def _read_web_crawl(
    path: Path,
    registered: dict[str, Any],
    query: str,
    preferred_chunk_ids: set[str],
) -> dict[str, Any]:
    chunks = load_source_chunks(registered)
    settings = load_web_fetch_settings()
    selected = relevant_chunks(
        query,
        chunks,
        top_k=settings.relevant_chunks_per_source,
        preferred_chunk_ids=preferred_chunk_ids,
    )
    selected = [
        {
            **chunk,
            "text": relevant_excerpt(
                query,
                str(chunk.get("text") or ""),
                max_chars=settings.relevant_excerpt_chars,
            ),
        }
        for chunk in selected
    ]
    return {
        **_base_record(path, registered),
        "columns": [],
        "sample_rows": [],
        "chunks": selected,
        "metadata": {
            "character_count": int(registered.get("character_count") or 0),
            "chunk_count": len(chunks),
            "returned_chunk_count": len(selected),
            "content_sha256": str(registered.get("content_sha256") or ""),
            "size_bytes": path.stat().st_size,
        },
    }


def _read_one(path: Path, *, query: str, preferred_chunk_ids: set[str]) -> dict[str, Any]:
    if not path.is_file():
        raise ValueError("path is not a file")
    registered = _registered_source(path)
    if str(registered.get("source_kind") or "") == "web_crawl":
        return _read_web_crawl(path, registered, query, preferred_chunk_ids)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _read_tabular(path, registered, ",")
    if suffix == ".tsv":
        return _read_tabular(path, registered, "\t")
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(path, registered)
    if suffix == ".xls":
        raise ValueError("legacy .xls is unsupported; save the file as .xlsx and upload again")
    if suffix == ".json":
        return _read_json(path, registered)
    if suffix in {".txt", ".md", ".html", ".htm", ".py"}:
        return _text_record(path, registered)
    raise ValueError(f"unsupported source type: {suffix or '<none>'}")


def _truncate_source_record(record: dict[str, Any], remaining_chars: int) -> dict[str, Any]:
    budget = min(remaining_chars, MAX_CHARS_PER_SOURCE)
    if budget <= 0:
        return {
            **{key: value for key, value in record.items() if key not in {"chunks", "sample_rows"}},
            "chunks": [],
            "sample_rows": [],
            "metadata": {
                **dict(record.get("metadata") or {}),
                "truncated": True,
                "omitted_for_batch_budget": True,
            },
        }
    chunks = list(record.get("chunks") or [])
    sample_rows = list(record.get("sample_rows") or [])
    if chunks:
        kept: list[dict[str, Any]] = []
        used = 0
        for chunk in chunks[:MAX_CHUNKS_PER_SOURCE]:
            text = str(chunk.get("text") or "")
            if used >= budget:
                break
            if used + len(text) <= budget:
                kept.append(chunk)
                used += len(text)
                continue
            kept.append({**chunk, "text": text[: max(0, budget - used)]})
            used = budget
            break
        truncated = dict(record)
        truncated["chunks"] = kept
        truncated["sample_rows"] = []
        truncated["metadata"] = {
            **dict(record.get("metadata") or {}),
            "truncated": len(kept) < len(chunks) or used >= budget,
            "returned_chunk_count": len(kept),
            "returned_character_count": used,
        }
        return truncated
    if sample_rows:
        kept_rows: list[dict[str, Any]] = []
        used = 0
        for row in sample_rows[:MAX_PREVIEW_ROWS]:
            encoded = json.dumps(row, ensure_ascii=False)
            if used + len(encoded) > budget and kept_rows:
                break
            kept_rows.append(row)
            used += len(encoded)
        truncated = dict(record)
        truncated["sample_rows"] = kept_rows
        truncated["chunks"] = []
        truncated["metadata"] = {
            **dict(record.get("metadata") or {}),
            "truncated": len(kept_rows) < len(sample_rows),
            "returned_row_count": len(kept_rows),
            "returned_character_count": used,
        }
        return truncated
    return record


def _assigned_source_paths() -> list[str]:
    try:
        from knowcoder_workspace_builder.runtime.invocation_context import (
            active_invocation_context,
        )

        context = active_invocation_context()
    except Exception:
        return []
    paths: list[str] = []
    for item in context.input.get("sources") or []:
        if not isinstance(item, dict):
            continue
        file_path = str(item.get("file_path") or "").strip()
        if file_path:
            paths.append(file_path)
    return paths


def _resolve_requested_paths(file_paths: list[str]) -> list[str]:
    tokens = [str(item).strip() for item in file_paths if str(item).strip()]
    if not tokens:
        return []
    if (
        any(token == ASSIGNED_SOURCES_TOKEN for token in tokens)
        or len(tokens) == 1
        and tokens[0].casefold()
        in {
            "all",
            "assigned",
            "assigned_sources",
        }
    ):
        assigned = _assigned_source_paths()
        return assigned or tokens
    return tokens


def _reader_query(question: str) -> str:
    context = active_invocation_context()
    workspace_context = context.input.get("workspace_context")
    units = workspace_context.get("extraction_units") if isinstance(workspace_context, dict) else None
    current = units[0] if isinstance(units, list) and units and isinstance(units[0], dict) else {}
    parts = [
        str(current.get("step") or "").strip(),
        " ".join(str(item).strip() for item in current.get("requirements") or [] if str(item).strip()),
        str(question or "").strip(),
    ]
    return " ".join(part for part in parts if part)


def _preferred_chunks_by_source() -> dict[str, set[str]]:
    context = active_invocation_context()
    workspace_context = context.input.get("workspace_context")
    units = workspace_context.get("extraction_units") if isinstance(workspace_context, dict) else None
    current = units[0] if isinstance(units, list) and units and isinstance(units[0], dict) else {}
    result: dict[str, set[str]] = {}
    for ref in current.get("chunk_refs") or []:
        if not isinstance(ref, dict):
            continue
        source_id = str(ref.get("source_id") or "").strip()
        chunk_id = str(ref.get("chunk_id") or "").strip()
        if source_id and chunk_id:
            result.setdefault(source_id, set()).add(chunk_id)
    return result


def _assigned_chunks(*, stage: str, units_field: str) -> list[dict[str, Any]]:
    """Resolve a stage-owned source assignment into one model input per chunk."""
    context = active_invocation_context()
    if context.stage != stage:
        raise ValueError(f"assigned chunks are available only during the {stage} stage")
    workspace_context = context.input.get("workspace_context")
    units = workspace_context.get(units_field) if isinstance(workspace_context, dict) else None
    if not isinstance(units, list) or not units:
        raise ValueError(f"{stage} stage input is missing {units_field}")

    source_by_id = {
        str(item.get("source_id") or "").strip(): item
        for item in context.input.get("sources") or []
        if isinstance(item, dict) and str(item.get("source_id") or "").strip()
    }
    if not source_by_id:
        raise ValueError(f"{stage} stage input has no assigned sources")

    resolved: list[dict[str, Any]] = []
    selected_refs: list[dict[str, str]] = []
    resolution_errors: list[dict[str, Any]] = []
    consecutive_failures = 0

    def record_failure(unit_index: int, source_id: str, chunk_id: str, error: Exception) -> None:
        nonlocal consecutive_failures
        consecutive_failures += 1
        resolution_errors.append(
            {
                "unit_index": unit_index,
                "source_id": source_id,
                "chunk_id": chunk_id,
                "error_type": type(error).__name__,
                "error": str(error),
            }
        )
        if consecutive_failures >= CONSECUTIVE_FAILURE_LIMIT:
            raise ValueError(
                f"{stage} source resolution failed for {CONSECUTIVE_FAILURE_LIMIT} consecutive inputs"
            ) from error

    def record_success(item: dict[str, Any]) -> None:
        nonlocal consecutive_failures
        resolved.append(item)
        consecutive_failures = 0
        chunk_id = str(item.get("chunk_id") or "")
        if chunk_id:
            selected_refs.append({"source_id": str(item["source_id"]), "chunk_id": chunk_id})

    for fallback_index, raw_unit in enumerate(units, start=1):
        if not isinstance(raw_unit, dict):
            record_failure(
                fallback_index,
                "",
                "",
                ValueError(f"{units_field}[{fallback_index}] must be an object"),
            )
            continue
        try:
            unit_index = int(raw_unit.get("unit_index") or fallback_index)
            step_index = int(raw_unit.get("step_index") or 1)
        except (TypeError, ValueError) as error:
            record_failure(fallback_index, "", "", error)
            continue
        step = str(raw_unit.get("step") or "").strip()
        requirements = [str(item).strip() for item in raw_unit.get("requirements") or [] if str(item).strip()]
        source_ids = [str(item).strip() for item in raw_unit.get("source_ids") or [] if str(item).strip()]
        if len(source_ids) != 1 or source_ids[0] not in source_by_id:
            record_failure(
                unit_index,
                source_ids[0] if len(source_ids) == 1 else "",
                "",
                ValueError(f"{units_field} unit {unit_index} must reference exactly one assigned source"),
            )
            continue
        source_id = source_ids[0]
        source = source_by_id[source_id]
        file_path = str(source.get("file_path") or "").strip()
        if not file_path:
            record_failure(unit_index, source_id, "", ValueError(f"assigned source {source_id} is missing file_path"))
            continue
        try:
            path = resolve_path(file_path)
            registered = _registered_source(path)
        except Exception as error:
            record_failure(unit_index, source_id, "", error)
            continue
        refs = [item for item in raw_unit.get("chunk_refs") or [] if isinstance(item, dict)]

        if refs:
            try:
                chunks_by_id = {
                    str(item.get("chunk_id") or "").strip(): item
                    for item in load_source_chunks(registered)
                    if str(item.get("chunk_id") or "").strip()
                }
            except Exception as error:
                record_failure(unit_index, source_id, "", error)
                continue
            for ref in refs:
                ref_source_id = str(ref.get("source_id") or "").strip()
                chunk_id = str(ref.get("chunk_id") or "").strip()
                if ref_source_id != source_id or not chunk_id:
                    record_failure(
                        unit_index,
                        source_id,
                        chunk_id,
                        ValueError(f"{units_field} unit {unit_index} contains an invalid chunk reference"),
                    )
                    continue
                chunk = chunks_by_id.get(chunk_id)
                if chunk is None:
                    record_failure(
                        unit_index,
                        source_id,
                        chunk_id,
                        ValueError(f"assigned chunk does not exist: {source_id}/{chunk_id}"),
                    )
                    continue
                text = str(chunk.get("text") or "").strip()
                if not text:
                    record_failure(
                        unit_index,
                        source_id,
                        chunk_id,
                        ValueError(f"assigned chunk is empty: {source_id}/{chunk_id}"),
                    )
                    continue
                record_success(
                    {
                        "unit_index": unit_index,
                        "step_index": step_index,
                        "step": step,
                        "requirements": requirements,
                        "source_id": source_id,
                        "chunk_id": chunk_id,
                        "title": str(source.get("title") or registered.get("title") or path.name),
                        "url": str(registered.get("url") or ""),
                        "text": text,
                    }
                )
            continue

        try:
            record = _read_one(path, query=" ".join([step, *requirements]), preferred_chunk_ids=set())
        except Exception as error:
            record_failure(unit_index, source_id, "", error)
            continue
        chunks = [item for item in record.get("chunks") or [] if isinstance(item, dict)]
        if chunks:
            for chunk_position, chunk in enumerate(chunks, start=1):
                chunk_id = str(chunk.get("chunk_id") or "").strip()
                text = str(chunk.get("text") or "").strip()
                if not chunk_id or not text:
                    record_failure(
                        unit_index,
                        source_id,
                        chunk_id,
                        ValueError(f"source {source_id} returned an invalid chunk at position {chunk_position}"),
                    )
                    continue
                record_success(
                    {
                        "unit_index": unit_index,
                        "step_index": step_index,
                        "step": step,
                        "requirements": requirements,
                        "source_id": source_id,
                        "chunk_id": chunk_id,
                        "title": str(record.get("title") or path.name),
                        "url": str(record.get("url") or ""),
                        "text": text,
                    }
                )
            continue

        sample_rows = record.get("sample_rows") or []
        if not isinstance(sample_rows, list) or not sample_rows:
            record_failure(
                unit_index,
                source_id,
                "",
                ValueError(f"assigned source has no readable extraction content: {source_id}"),
            )
            continue
        record_success(
            {
                "unit_index": unit_index,
                "step_index": step_index,
                "step": step,
                "requirements": requirements,
                "source_id": source_id,
                "chunk_id": "",
                "title": str(record.get("title") or path.name),
                "url": str(record.get("url") or ""),
                "text": json.dumps(sample_rows, ensure_ascii=False),
            }
        )

    if not resolved:
        raise ValueError(f"{stage} stage resolved no readable chunks")
    paths = active_session_paths()
    AtomicWriter(paths).json(
        paths.attempts / context.attempt_id / "source_resolution_errors.json",
        {"format_version": 1, "count": len(resolution_errors), "items": resolution_errors},
    )
    AtomicWriter(paths).json(
        paths.attempts / context.attempt_id / "selected_chunks.json",
        {"format_version": 1, "evidence_refs": selected_refs},
    )
    return resolved


def assigned_extraction_chunks() -> list[dict[str, Any]]:
    """Resolve extraction assignments and split Schema chunks into smaller model inputs."""
    parent_chunks = _assigned_chunks(stage="extract", units_field="extraction_units")
    settings = load_web_fetch_settings()
    chunks: list[dict[str, Any]] = []
    evidence_refs: list[dict[str, str]] = []
    for parent in parent_chunks:
        parent_chunk_id = str(parent.get("chunk_id") or "")
        if parent_chunk_id:
            evidence_refs.append({"source_id": str(parent["source_id"]), "chunk_id": parent_chunk_id})
        parts = token_chunks(
            str(parent["text"]),
            target_tokens=settings.extraction_chunk_target_tokens,
            overlap_tokens=settings.extraction_chunk_overlap_tokens,
        )
        for part_index, part in enumerate(parts, start=1):
            chunks.append(
                {
                    **parent,
                    "chunk_id": (
                        f"{parent_chunk_id}#extract_{part_index:04d}"
                        if parent_chunk_id
                        else f"{parent['source_id']}#extract_{len(chunks) + 1:04d}"
                    ),
                    "evidence_chunk_id": parent_chunk_id,
                    "text": part["text"],
                    "token_count": part["token_count"],
                    "tokenizer_model": part["tokenizer_model"],
                }
            )
    paths = active_session_paths()
    context = active_invocation_context()
    AtomicWriter(paths).json(
        paths.attempts / context.attempt_id / "selected_chunks.json",
        {"format_version": 1, "evidence_refs": list({(item["source_id"], item["chunk_id"]): item for item in evidence_refs}.values())},
    )
    return chunks


def assigned_schema_chunks() -> list[dict[str, Any]]:
    """Resolve the Schema assignment into one model input per evidence chunk."""
    chunks = _assigned_chunks(stage="schema_build", units_field="schema_units")
    unique: dict[tuple[str, str], dict[str, Any]] = {}
    for chunk in chunks:
        key = (str(chunk.get("source_id") or ""), str(chunk.get("chunk_id") or ""))
        unique.setdefault(key, chunk)
    return list(unique.values())


def _persist_selected_chunks(sources: list[dict[str, Any]]) -> None:
    context = active_invocation_context()
    if context.stage != "extract":
        return
    paths = active_session_paths()
    refs = [
        {"source_id": str(source.get("source_id") or ""), "chunk_id": str(chunk.get("chunk_id") or "")}
        for source in sources
        for chunk in source.get("chunks") or []
        if str(source.get("source_id") or "").strip() and str(chunk.get("chunk_id") or "").strip()
    ]
    AtomicWriter(paths).json(
        paths.attempts / context.attempt_id / "selected_chunks.json",
        {"format_version": 1, "evidence_refs": refs},
    )


@tool
def source_reader(file_paths: list[str], question: str = "") -> str:
    """Read exact registered current-Session files and return structured source data.

    Args:
        file_paths: Absolute current-Session virtual paths supplied by the Coordinator.
            Pass `["*"]` during extraction to read every assigned source in one call.
        question: Current question, used only as caller context and never persisted as evidence.
    """
    if not isinstance(file_paths, list) or not file_paths:
        return json.dumps({"ok": False, "sources": [], "errors": [{"error": "file_paths must be non-empty"}]})
    requested = _resolve_requested_paths(file_paths)
    if not requested:
        return json.dumps({"ok": False, "sources": [], "errors": [{"error": "file_paths must be non-empty"}]})
    context = active_invocation_context()
    if context.stage == "extract" and len(requested) > MAX_UNSTRUCTURED_SOURCE_BATCH:
        return json.dumps(
            {
                "ok": False,
                "error_type": "source_batch_too_large",
                "error": "Unstructured extraction source batch exceeds the supported size.",
                "max_batch_sources": MAX_UNSTRUCTURED_SOURCE_BATCH,
                "requested_sources": len(requested),
                "repair_instruction": (
                    "Read the next assigned sources in a batch within max_batch_sources. "
                    "Persist that batch before reading the next batch."
                ),
            },
            ensure_ascii=False,
        )
    if len(requested) > MAX_BATCH_SOURCES:
        requested = requested[:MAX_BATCH_SOURCES]
    sources: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    remaining_chars = MAX_BATCH_TEXT_CHARS
    query = _reader_query(question)
    preferred_by_source = _preferred_chunks_by_source()
    for value in requested:
        try:
            path = resolve_path(value)
            registered = _registered_source(path)
            source_id = str(registered.get("source_id") or "")
            record = _read_one(
                path,
                query=query,
                preferred_chunk_ids=preferred_by_source.get(source_id, set()),
            )
            if record.get("chunks") or record.get("sample_rows"):
                record = _truncate_source_record(record, remaining_chars)
                used = int((record.get("metadata") or {}).get("returned_character_count") or 0)
                if not used:
                    used = sum(len(str(chunk.get("text") or "")) for chunk in record.get("chunks") or [])
                    used += sum(len(json.dumps(row, ensure_ascii=False)) for row in record.get("sample_rows") or [])
                remaining_chars = max(0, remaining_chars - used)
            sources.append(record)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            errors.append({"path": str(value), "error": str(exc)})
    _persist_selected_chunks(sources)
    return json.dumps(
        {
            "ok": not errors,
            "sources": sources,
            "errors": errors,
            "batch": {
                "requested": len(requested),
                "returned": len(sources),
                "char_budget": MAX_BATCH_TEXT_CHARS,
                "remaining_char_budget": remaining_chars,
            },
        },
        ensure_ascii=False,
    )
